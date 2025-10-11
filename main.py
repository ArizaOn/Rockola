#!/usr/bin/env python3
from fastapi import FastAPI, Form, File, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from yt_dlp import YoutubeDL
from io import BytesIO
import os
import uuid
import shutil
import re
import openpyxl
import pandas as pd
import csv
from fastapi.staticfiles import StaticFiles
import subprocess
import sys
import time
import glob
from starlette.background import BackgroundTask

def cleanup_file(path, delay=5):
    """Limpia el archivo después de un delay para permitir la descarga"""
    time.sleep(delay)
    try:
        if os.path.exists(path):
            os.remove(path)
            print(f"✅ Archivo limpiado: {path}")
    except Exception as e:
        print(f"⚠️ No se pudo limpiar {path}: {e}")

# ============================================================
# 🛠️ LIMPIAR ARCHIVOS ANTIGUOS AL INICIAR
# ============================================================
def cleanup_old_files():
    """Limpia archivos de descargas anteriores"""
    try:
        old_folders = glob.glob("downloads/*")
        for folder in old_folders:
            try:
                if os.path.isdir(folder):
                    shutil.rmtree(folder)
                elif os.path.isfile(folder):
                    os.remove(folder)
            except Exception as e:
                print(f"⚠️ No se pudo limpiar {folder}: {e}")
        print("✅ Carpeta de descargas limpiada")
    except Exception as e:
        print(f"⚠️ Error en limpieza inicial: {e}")

# Ejecutar limpieza al iniciar
cleanup_old_files()

# Configurar FFmpeg
ffmpeg_dir = os.path.dirname(os.path.abspath(__file__))
os.environ['PATH'] = ffmpeg_dir + os.pathsep + os.environ['PATH']
os.environ['FFMPEG_BINARY'] = os.path.join(ffmpeg_dir, 'ffmpeg.exe')

# Verificar que FFmpeg existe
if not os.path.exists(os.path.join(ffmpeg_dir, 'ffmpeg.exe')):
    print("⚠️ ADVERTENCIA: ffmpeg.exe no encontrado en la carpeta raíz")

ydl_opts = {
    'quiet': True,
    'no_warnings': True,
    'ignoreerrors': True,
    'extractor_args': {'youtubetab': ['skip=authcheck']},
    'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
    'http_headers': {
        'Accept-Language': 'es-ES,es;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Referer': 'https://www.youtube.com/',
    },
}

# ============================================================
# 🚀 INICIAR FASTAPI
# ============================================================

app = FastAPI()

# Montar archivos estáticos
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("index.html")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# 🔧 CONFIGURACIÓN DE COOKIES
# ============================================================

COOKIES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookies.txt")
USE_COOKIES = os.path.exists(COOKIES_PATH) and os.path.getsize(COOKIES_PATH) > 100

print("=" * 50)
if USE_COOKIES:
    ydl_opts['cookiefile'] = COOKIES_PATH
    print(f"✅ Usando cookies desde: {COOKIES_PATH}")
else:
    print("⚠️ No se encontraron cookies. El servidor funcionará sin cookies.")
print("=" * 50)

# ============================================================
# 🛠️ FUNCIONES AUXILIARES
# ============================================================

def is_url(text: str) -> bool:
    """Verifica si el texto es una URL válida"""
    url_pattern = re.compile(
        r'^https?://'
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'
        r'localhost|'
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'
        r'(?::\d+)?'
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    return url_pattern.match(text) is not None


def process_excel_file(file_content: bytes) -> list:
    """Procesa archivo Excel de Exportify y extrae canciones"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        songs = []
        headers = {}

        for col_idx, cell in enumerate(sheet[1], start=1):
            header = str(cell.value).strip().lower() if cell.value else ""
            if 'track name' in header or 'song' in header or 'title' in header:
                headers['track'] = col_idx
            elif 'artist' in header:
                headers['artist'] = col_idx

        if 'track' not in headers:
            print("⚠️ No se encontró columna de nombre de canción")
            return []

        for row_idx in range(2, sheet.max_row + 1):
            track_name = sheet.cell(row=row_idx, column=headers.get('track', 1)).value
            artist_name = sheet.cell(row=row_idx, column=headers.get('artist', 2)).value if 'artist' in headers else ""

            if track_name:
                track_name = str(track_name).strip()
                artist_name = str(artist_name).strip() if artist_name else ""
                query = f"{track_name} {artist_name}".strip()
                songs.append(query)
                print(f"🎵 Extraído: {query}")

        print(f"\n✅ Total de canciones extraídas: {len(songs)}")
        return songs

    except Exception as e:
        print(f"❌ Error procesando Excel: {str(e)}")
        return []


def search_and_download(query: str, ydl_opts: dict) -> bool:
    """Busca una canción en YouTube y descarga la primera con audio"""
    try:
        search_url = f"ytsearch1:{query}"
        search_opts = {
            'quiet': True,
            'no_warnings': True,
            'extract_flat': False,
            'format': 'bestaudio/best',
        }

        if USE_COOKIES:
            search_opts['cookiefile'] = COOKIES_PATH

        with YoutubeDL(search_opts) as ydl:
            search_result = ydl.extract_info(search_url, download=False)

            if not search_result or 'entries' not in search_result:
                print(f"❌ No se encontraron resultados para: {query}")
                return False

            for entry in search_result['entries']:
                if entry and entry.get('id'):
                    try:
                        video_url = f"https://www.youtube.com/watch?v={entry['id']}"
                        if USE_COOKIES:
                            ydl_opts['cookiefile'] = COOKIES_PATH
                        with YoutubeDL(ydl_opts) as ydl_download:
                            ydl_download.download([video_url])
                        print(f"✅ Descargado: {query} -> {entry.get('title', 'Unknown')}")
                        return True
                    except Exception as e:
                        print(f"⚠️ Error con video {entry['id']}: {str(e)}")
                        continue

            print(f"❌ No se pudo descargar ningún resultado para: {query}")
            return False

    except Exception as e:
        print(f"❌ Error buscando '{query}': {str(e)}")
        return False


# ============================================================
# 📥 ENDPOINTS
# ============================================================

@app.post("/download/")
def download(url: str = Form(...), format_type: str = Form("mp3")):
    """Descarga individual desde URL"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)
    filename = str(uuid.uuid4())

    ydl_opts_download = {
        'quiet': True,
        'no_warnings': True,
    }

    if USE_COOKIES:
        ydl_opts_download['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts_download.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts_download.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'merge_output_format': 'mp4',
        })

    try:
        with YoutubeDL(ydl_opts_download) as ydl:
            ydl.download([url])

        # Esperar a que FFmpeg termine de procesar
        timeout = 120  # 2 minutos máximo
        start_time = time.time()
        file_found = None
        
        while time.time() - start_time < timeout:
            files_in_folder = os.listdir(output_folder)
            
            # Buscar archivos que coincidan con el UUID y NO terminen en _
            for file in files_in_folder:
                if file.startswith(filename) and not file.endswith('_'):
                    file_found = file
                    break
            
            if file_found:
                break
            
            time.sleep(0.5)
        
        if not file_found:
            print(f"❌ No se encontró archivo después de {timeout} segundos")
            print(f"Archivos en la carpeta: {os.listdir(output_folder)}")
            return {"error": "Timeout: No se pudo completar la conversión del archivo"}
        
        file_path = os.path.join(output_folder, file_found)
        
        print(f"✅ Archivo encontrado: {file_found}")
        
        return FileResponse(
            path=file_path,
            filename=file_found,
            media_type="application/octet-stream",
            background=BackgroundTask(cleanup_file, file_path, delay=10)
        )

    except Exception as e:
        error_msg = str(e)
        if "HTTP Error 403" in error_msg:
            error_msg = "Error 403: El vídeo podría tener restricciones de región o edad"
        elif "Video unavailable" in error_msg:
            error_msg = "Vídeo no disponible o URL inválida"
        elif "No video formats found" in error_msg:
            error_msg = "No se encontraron formatos disponibles para este vídeo"

        return {"error": error_msg}


@app.post("/download_batch/")
async def download_batch(file: UploadFile = File(...), format_type: str = Form("mp3")):
    """Descarga múltiple desde archivo .txt, .csv, .xlsx, .xls"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    content = await file.read()
    lines = []

    # Procesar según tipo de archivo
    if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
        print("📊 Procesando archivo Excel...")
        lines = process_excel_file(content)
    else:
        # Procesar archivos de texto
        lines = [line.strip() for line in content.decode('utf-8').strip().split('\n') if line.strip()]

    if not lines:
        return {"error": "El archivo está vacío o no se pudieron extraer canciones"}

    batch_id = str(uuid.uuid4())
    batch_folder = os.path.join(output_folder, batch_id)
    os.makedirs(batch_folder, exist_ok=True)

    ydl_opts_batch = {
        'quiet': True,
        'no_warnings': True,
        'ignoreerrors': True,
        'outtmpl': f'{batch_folder}/%(title)s.%(ext)s',
    }

    if USE_COOKIES:
        ydl_opts_batch['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts_batch.update({
            'format': 'bestaudio/best',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts_batch.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'merge_output_format': 'mp4',
        })

    zip_path = None
    successful_downloads = 0
    failed_downloads = []

    try:
        print(f"\n🎵 Iniciando descarga de {len(lines)} canciones...\n")

        for idx, line in enumerate(lines, 1):
            print(f"[{idx}/{len(lines)}] Procesando: {line}")

            if is_url(line):
                try:
                    with YoutubeDL(ydl_opts_batch) as ydl:
                        ydl.download([line])
                    successful_downloads += 1
                    print(f"✅ Descargado desde URL")
                except Exception as e:
                    print(f"❌ Error con URL: {str(e)}")
                    failed_downloads.append(line)
            else:
                if search_and_download(line, ydl_opts_batch):
                    successful_downloads += 1
                else:
                    failed_downloads.append(line)

        if not os.listdir(batch_folder) or successful_downloads == 0:
            shutil.rmtree(batch_folder)
            error_msg = "No se pudo descargar ningún archivo."
            if failed_downloads:
                error_msg += f" Fallaron: {', '.join(failed_downloads[:5])}"
            return {"error": error_msg}

        zip_path = f"{output_folder}/{batch_id}"
        shutil.make_archive(zip_path, 'zip', batch_folder)
        shutil.rmtree(batch_folder)

        print(f"\n📦 Completado: {successful_downloads} exitosas, {len(failed_downloads)} fallidas")

        return FileResponse(
            path=f"{zip_path}.zip",
            filename="batch_download.zip",
            media_type="application/zip",
            background=BackgroundTask(cleanup_file, f"{zip_path}.zip", delay=10)
        )
    except Exception as e:
        if os.path.exists(batch_folder):
            shutil.rmtree(batch_folder)
        if zip_path and os.path.exists(f"{zip_path}.zip"):
            os.remove(f"{zip_path}.zip")
        return {"error": f"Error al procesar descargas: {str(e)}"}


@app.post("/download_playlist/")
def download_playlist(url: str = Form(...), format_type: str = Form("mp3")):
    """Descarga playlist completa"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    playlist_id = str(uuid.uuid4())
    playlist_folder = os.path.join(output_folder, playlist_id)
    os.makedirs(playlist_folder, exist_ok=True)

    ydl_opts_playlist = {
        'quiet': True,
        'no_warnings': True,
        'ignoreerrors': True,
    }

    if USE_COOKIES:
        ydl_opts_playlist['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts_playlist.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts_playlist.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'merge_output_format': 'mp4',
        })

    zip_path = None
    try:
        with YoutubeDL(ydl_opts_playlist) as ydl:
            ydl.download([url])

        if not os.listdir(playlist_folder):
            shutil.rmtree(playlist_folder)
            return {"error": "No se pudo descargar ningún vídeo de la playlist"}

        zip_path = f"{output_folder}/{playlist_id}"
        shutil.make_archive(zip_path, 'zip', playlist_folder)
        shutil.rmtree(playlist_folder)

        print(f"📦 Playlist descargada exitosamente")

        return FileResponse(
            path=f"{zip_path}.zip",
            filename="playlist_download.zip",
            media_type="application/zip",
            background=BackgroundTask(cleanup_file, f"{zip_path}.zip")
        )
    except Exception as e:
        if os.path.exists(playlist_folder):
            shutil.rmtree(playlist_folder)
        if zip_path and os.path.exists(f"{zip_path}.zip"):
            os.remove(f"{zip_path}.zip")

        error_msg = str(e)
        if "playlist" in error_msg.lower() and "not" in error_msg.lower():
            error_msg = "La URL no parece ser una playlist válida"

        return {"error": error_msg}


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)