from fastapi import FastAPI, Form, File, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from yt_dlp import YoutubeDL
import os
import uuid
import shutil
import re
import openpyxl
import pandas as pd
import csv

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------- CONFIG --------------------
COOKIES_FILE = "cookies.txt"  # <-- tu archivo de cookies exportado desde el navegador

# -------------------- UTILIDADES --------------------
def is_url(text: str) -> bool:
    """Verifica si el texto es una URL válida"""
    url_pattern = re.compile(
        r'^https?://'  # http:// o https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'  # dominio
        r'localhost|'  # localhost
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # IP
        r'(?::\d+)?'  # puerto opcional
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    return url_pattern.match(text) is not None


def load_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xls", ".xlsx"]:
        return pd.read_excel(path, dtype=str, engine='openpyxl', keep_default_na=False)
    elif ext in [".csv", ".txt"]:
        with open(path, "rb") as fb:
            sample = fb.read(32768)
        encodings_to_try = ["utf-8", "latin-1", "cp1252"]
        last_err = None
        for enc in encodings_to_try:
            try:
                s = sample.decode(enc)
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(s)
                delim = dialect.delimiter
                df = pd.read_csv(path, dtype=str, encoding=enc, sep=delim, keep_default_na=False)
                return df
            except Exception as e:
                last_err = e
                continue
        raise ValueError(f"No pude leer el CSV. Último error: {last_err}")
    else:
        raise ValueError("Formato no soportado. Usa .csv, .txt, .xls o .xlsx")


def process_excel_file(file_content: bytes) -> list:
    """Procesa archivo Excel y extrae canciones"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active
        songs = []
        headers = {}

        for col_idx, cell in enumerate(sheet[1], start=1):
            header = str(cell.value).strip().lower() if cell.value else ""
            if 'track name' in header or 'song' in header or 'title' in header:
                headers['track'] = col_idx
            elif 'artist' in header:
                headers['artist'] = col_idx

        if 'track' not in headers:
            print("⚠️ No se encontró columna de nombre de canción")
            return []

        for row_idx in range(2, sheet.max_row + 1):
            track_name = sheet.cell(row=row_idx, column=headers.get('track', 1)).value
            artist_name = sheet.cell(row=row_idx, column=headers.get('artist', 2)).value if 'artist' in headers else ""

            if track_name:
                track_name = str(track_name).strip()
                artist_name = str(artist_name).strip() if artist_name else ""
                query = f"{track_name} {artist_name}".strip()
                songs.append(query)
                print(f"📝 Extraído: {query}")

        print(f"\n✅ Total de canciones extraídas: {len(songs)}")
        return songs

    except Exception as e:
        print(f"❌ Error procesando Excel: {str(e)}")
        return []


# -------------------- DESCARGA INDIVIDUAL --------------------
@app.post("/download/")
def download(url: str = Form(...), format_type: str = Form("mp3")):
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)
    filename = str(uuid.uuid4())

    ydl_opts = {
        'quiet': False,
        'no_warnings': False,
        'cookiefile': COOKIES_FILE,
        'extractor_args': {'youtubetab': ['skip=authcheck']},
    }

    if format_type == "mp3":
        ydl_opts.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'postprocessors': [{'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3', 'preferredquality': '192'}],
        })
    else:
        ydl_opts.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'merge_output_format': 'mp4',
        })

    try:
        with YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])

        for file in os.listdir(output_folder):
            if file.startswith(filename):
                file_path = os.path.join(output_folder, file)
                def cleanup():
                    try: os.remove(file_path)
                    except Exception: pass
                return FileResponse(path=file_path, filename=file, media_type="application/octet-stream", background=cleanup)

        return {"error": "No se pudo encontrar el archivo descargado"}

    except Exception as e:
        return {"error": str(e)}


# -------------------- DESCARGA BATCH --------------------
@app.post("/download_batch/")
async def download_batch(file: UploadFile = File(...), format_type: str = Form("mp3")):
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    content = await file.read()
    lines = [line.strip() for line in content.decode('utf-8').strip().split('\n') if line.strip()]

    if not lines:
        return {"error": "El archivo está vacío"}

    batch_id = str(uuid.uuid4())
    batch_folder = os.path.join(output_folder, batch_id)
    os.makedirs(batch_folder, exist_ok=True)

    successful_downloads = 0
    failed_downloads = []

    for i, line in enumerate(lines, 1):
        if is_url(line):
            search_query = line
        else:
            search_query = f"ytsearch1:{line}"

        ydl_opts = {
            'quiet': False,
            'no_warnings': False,
            'ignoreerrors': True,
            'cookiefile': COOKIES_FILE,
            'extractor_args': {'youtubetab': ['skip=authcheck']},
            'outtmpl': f'{batch_folder}/%(title)s.%(ext)s'
        }

        if format_type == "mp3":
            ydl_opts['postprocessors'] = [{'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3', 'preferredquality': '192'}]
        else:
            ydl_opts['merge_output_format'] = 'mp4'

        try:
            with YoutubeDL(ydl_opts) as ydl:
                ydl.extract_info(search_query, download=True)
            successful_downloads += 1
        except Exception:
            failed_downloads.append(line)

    if not os.listdir(batch_folder) or successful_downloads == 0:
        shutil.rmtree(batch_folder)
        return {"error": "No se pudo descargar ningún archivo"}

    zip_path = f"{output_folder}/{batch_id}"
    shutil.make_archive(zip_path, 'zip', batch_folder)
    shutil.rmtree(batch_folder)

    def cleanup():
        try: os.remove(f"{zip_path}.zip")
        except Exception: pass

    return FileResponse(path=f"{zip_path}.zip", filename="batch_download.zip", media_type="application/zip", background=cleanup)


# -------------------- DESCARGA PLAYLIST --------------------
@app.post("/download_playlist/")
def download_playlist(url: str = Form(...), format_type: str = Form("mp3")):
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    playlist_id = str(uuid.uuid4())
    playlist_folder = os.path.join(output_folder, playlist_id)
    os.makedirs(playlist_folder, exist_ok=True)

    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
        'ignoreerrors': True,
        'cookiefile': COOKIES_FILE,
        'extractor_args': {'youtubetab': ['skip=authcheck']},
    }

    if format_type == "mp3":
        ydl_opts.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'postprocessors': [{'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3', 'preferredquality': '192'}],
        })
    else:
        ydl_opts.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'merge_output_format': 'mp4',
        })

    try:
        with YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])

        if not os.listdir(playlist_folder):
            shutil.rmtree(playlist_folder)
            return {"error": "No se pudo descargar ningún video de la playlist"}

        zip_path = f"{output_folder}/{playlist_id}"
        shutil.make_archive(zip_path, 'zip', playlist_folder)
        shutil.rmtree(playlist_folder)

        def cleanup():
            try: os.remove(f"{zip_path}.zip")
            except Exception: pass

        return FileResponse(path=f"{zip_path}.zip", filename="playlist_download.zip", media_type="application/zip", background=cleanup)

    except Exception as e:
        if os.path.exists(playlist_folder):
            shutil.rmtree(playlist_folder)
        return {"error": str(e)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
