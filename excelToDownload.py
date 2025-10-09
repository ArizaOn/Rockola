#!/usr/bin/env python3
from fastapi import FastAPI, Form, File, UploadFile
from fastapi.responses import FileResponse
from io import BytesIO
from fastapi.middleware.cors import CORSMiddleware
from yt_dlp import YoutubeDL
import os
import uuid
import shutil
import openpyxl
import pandas as pd
import csv
import subprocess
import sys
from fastapi.staticfiles import StaticFiles

# ============================================================
# 🔧 GENERADOR DE COOKIES - Ejecuta al iniciar
# ============================================================

COOKIES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookies.txt")

def generate_cookies():
    """Genera el archivo cookies.txt usando yt-dlp y Chrome"""
    
    print("🔄 Generando cookies desde Chrome...")
    print(f"📁 Ruta: {COOKIES_PATH}")
    
    try:
        # Comando para generar cookies
        cmd = [
            "yt-dlp",
            "--cookies-from-browser", "chrome",
            "--write-cookies",
            "--cookies", COOKIES_PATH,
            "https://www.youtube.com"
        ]
        
        # Ejecutar el comando
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0:
            if os.path.exists(COOKIES_PATH):
                file_size = os.path.getsize(COOKIES_PATH)
                print(f"✅ Cookies generadas exitosamente!")
                print(f"📊 Tamaño del archivo: {file_size} bytes")
                return True
            else:
                print("⚠️ El archivo de cookies no se creó")
                return False
        else:
            print(f"❌ Error al generar cookies:")
            print(f"STDOUT: {result.stdout}")
            print(f"STDERR: {result.stderr}")
            return False
            
    except subprocess.TimeoutExpired:
        print("❌ Timeout: El proceso tardó demasiado")
        return False
    except FileNotFoundError:
        print("❌ Error: yt-dlp no está instalado")
        print("Instálalo con: pip install yt-dlp")
        return False
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        return False

def check_cookies():
    """Verifica si el archivo de cookies existe y es válido"""
    if os.path.exists(COOKIES_PATH):
        file_size = os.path.getsize(COOKIES_PATH)
        if file_size > 100:  # Al menos 100 bytes
            print(f"✅ Archivo de cookies existente encontrado ({file_size} bytes)")
            return True
        else:
            print(f"⚠️ Archivo de cookies muy pequeño ({file_size} bytes), regenerando...")
            return False
    return False

# Ejecutar generador de cookies al iniciar
print("=" * 50)
print("🎬 Generador de Cookies para yt-dlp")
print("=" * 50)

if check_cookies():
    print("💾 Usando cookies existentes")
else:
    print("\n📝 Creando nuevas cookies...")
    if generate_cookies():
        print("\n✅ Cookies listas!")
    else:
        print("\n⚠️ No se pudo generar las cookies automáticamente")
        print("El servidor intentará funcionar sin cookies")

print("=" * 50)

# ============================================================
# 🚀 INICIAR FASTAPI
# ============================================================

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("index.html")

# ⚠️ IMPORTANTE: Configurar CORS para permitir peticiones desde el frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En producción, especifica el dominio exacto
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# 🔹 Ruta al archivo de cookies
COOKIES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookies.txt")
USE_COOKIES = os.path.exists(COOKIES_PATH)


@app.post("/download/")
def download(url: str = Form(...), format_type: str = Form("mp3")):
    """Descarga individual desde URL"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)
    filename = str(uuid.uuid4())

    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
    }

    # Si hay cookies disponibles, agrégalas
    if USE_COOKIES:
        ydl_opts['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{output_folder}/{filename}.%(ext)s',
            'merge_output_format': 'mp4',
        })

    try:
        with YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])

        for file in os.listdir(output_folder):
            if file.startswith(filename):
                file_path = os.path.join(output_folder, file)

                def cleanup():
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    except Exception:
                        pass

                return FileResponse(
                    path=file_path,
                    filename=file,
                    media_type="application/octet-stream",
                    background=cleanup
                )

        return {"error": "No se pudo encontrar el archivo descargado"}

    except Exception as e:
        error_msg = str(e)
        if "HTTP Error 403" in error_msg:
            error_msg = "Error 403: El video podría tener restricciones de región o edad"
        elif "Video unavailable" in error_msg:
            error_msg = "Video no disponible o URL inválida"
        elif "No video formats found" in error_msg:
            error_msg = "No se encontraron formatos disponibles para este video"

        return {"error": error_msg}


def is_url(text: str) -> bool:
    """Verifica si el texto es una URL válida"""
    return text.startswith(('http://', 'https://', 'www.'))


def search_and_download(query: str, ydl_opts: dict) -> bool:
    """Busca una canción en YouTube y descarga la primera con audio"""
    try:
        search_url = f"ytsearch3:{query}"
        search_opts = {
            'quiet': True,
            'no_warnings': True,
            'extract_flat': False,
            'format': 'bestaudio/best',
        }

        if USE_COOKIES:
            search_opts['cookiefile'] = COOKIES_PATH

        with YoutubeDL(search_opts) as ydl:
            search_result = ydl.extract_info(search_url, download=False)

            if not search_result or 'entries' not in search_result:
                print(f"❌ No se encontraron resultados para: {query}")
                return False

            for entry in search_result['entries']:
                if entry and entry.get('id'):
                    try:
                        video_url = f"https://www.youtube.com/watch?v={entry['id']}"
                        if USE_COOKIES:
                            ydl_opts['cookiefile'] = COOKIES_PATH
                        with YoutubeDL(ydl_opts) as ydl_download:
                            ydl_download.download([video_url])
                        print(f"✅ Descargado: {query} -> {entry.get('title', 'Unknown')}")
                        return True
                    except Exception as e:
                        print(f"⚠️ Error con video {entry['id']}: {str(e)}")
                        continue

            print(f"❌ No se pudo descargar ningún resultado para: {query}")
            return False

    except Exception as e:
        print(f"❌ Error buscando '{query}': {str(e)}")
        return False


def process_excel_file(file_content: bytes) -> list:
    """Procesa archivo Excel de Exportify y extrae canciones"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        songs = []
        headers = {}

        for col_idx, cell in enumerate(sheet[1], start=1):
            header = str(cell.value).strip().lower() if cell.value else ""
            if 'track name' in header or 'song' in header or 'title' in header:
                headers['track'] = col_idx
            elif 'artist' in header:
                headers['artist'] = col_idx

        if 'track' not in headers:
            print("⚠️ No se encontró columna de nombre de canción")
            return []

        for row_idx in range(2, sheet.max_row + 1):
            track_name = sheet.cell(row=row_idx, column=headers.get('track', 1)).value
            artist_name = sheet.cell(row=row_idx, column=headers.get('artist', 2)).value if 'artist' in headers else ""

            if track_name:
                track_name = str(track_name).strip()
                artist_name = str(artist_name).strip() if artist_name else ""
                query = f"{track_name} {artist_name}".strip()
                songs.append(query)
                print(f"📝 Extraído: {query}")

        print(f"\n✅ Total de canciones extraídas: {len(songs)}")
        return songs

    except Exception as e:
        print(f"❌ Error procesando Excel: {str(e)}")
        return []


def load_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xls", ".xlsx"]:
        return pd.read_excel(path, dtype=str, engine='openpyxl', keep_default_na=False)
    elif ext in [".csv", ".txt"]:
        with open(path, "rb") as fb:
            sample = fb.read(32768)
        encodings_to_try = ["utf-8", "latin-1", "cp1252"]
        last_err = None
        for enc in encodings_to_try:
            try:
                s = sample.decode(enc)
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(s)
                delim = dialect.delimiter
                df = pd.read_csv(path, dtype=str, encoding=enc, sep=delim, keep_default_na=False)
                return df
            except Exception as e:
                last_err = e
                continue
        raise ValueError(f"No pude leer el CSV. Último error: {last_err}")
    else:
        raise ValueError("Formato no soportado. Usa .csv, .txt, .xls o .xlsx")


@app.post("/download_batch/")
async def download_batch(file: UploadFile = File(...), format_type: str = Form("mp3")):
    """Descarga múltiple desde archivo .txt o .xlsx (URLs o nombres de canciones)"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    content = await file.read()
    lines = []

    if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
        print("📊 Procesando archivo Excel de Exportify...")
        lines = process_excel_file(content)
    else:
        lines = [line.strip() for line in content.decode('utf-8').strip().split('\n') if line.strip()]

    if not lines:
        return {"error": "El archivo está vacío o no se pudieron extraer canciones"}

    batch_id = str(uuid.uuid4())
    batch_folder = os.path.join(output_folder, batch_id)
    os.makedirs(batch_folder, exist_ok=True)

    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
        'ignoreerrors': True,
        'outtmpl': f'{batch_folder}/%(title)s.%(ext)s',
    }

    if USE_COOKIES:
        ydl_opts['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts.update({
            'format': 'bestaudio/best',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'merge_output_format': 'mp4',
        })

    zip_path = None
    successful_downloads = 0
    failed_downloads = []

    try:
        print(f"\n🎵 Iniciando descarga de {len(lines)} canciones...\n")

        for idx, line in enumerate(lines, 1):
            print(f"[{idx}/{len(lines)}] Procesando: {line}")

            if is_url(line):
                try:
                    with YoutubeDL(ydl_opts) as ydl:
                        ydl.download([line])
                    successful_downloads += 1
                    print(f"✅ Descargado desde URL: {line}")
                except Exception as e:
                    print(f"❌ Error con URL {line}: {str(e)}")
                    failed_downloads.append(line)
            else:
                if search_and_download(line, ydl_opts):
                    successful_downloads += 1
                else:
                    failed_downloads.append(line)

        if not os.listdir(batch_folder) or successful_downloads == 0:
            shutil.rmtree(batch_folder)
            error_msg = "No se pudo descargar ningún archivo."
            if failed_downloads:
                error_msg += f" Fallaron: {', '.join(failed_downloads[:5])}"
            return {"error": error_msg}

        zip_path = f"{output_folder}/{batch_id}"
        shutil.make_archive(zip_path, 'zip', batch_folder)
        shutil.rmtree(batch_folder)

        def cleanup():
            try:
                if os.path.exists(f"{zip_path}.zip"):
                    os.remove(f"{zip_path}.zip")
            except Exception:
                pass

        print(f"\n📦 Completado: {successful_downloads} exitosas, {len(failed_downloads)} fallidas")
        if failed_downloads:
            print(f"❌ Fallidas: {', '.join(failed_downloads[:10])}")

        return FileResponse(
            path=f"{zip_path}.zip",
            filename="batch_download.zip",
            media_type="application/zip",
            background=cleanup
        )
    except Exception as e:
        if os.path.exists(batch_folder):
            shutil.rmtree(batch_folder)
        if zip_path and os.path.exists(f"{zip_path}.zip"):
            os.remove(f"{zip_path}.zip")
        return {"error": f"Error al procesar descargas: {str(e)}"}


@app.post("/download_playlist/")
def download_playlist(url: str = Form(...), format_type: str = Form("mp3")):
    """Descarga playlist completa"""
    output_folder = "downloads"
    os.makedirs(output_folder, exist_ok=True)

    playlist_id = str(uuid.uuid4())
    playlist_folder = os.path.join(output_folder, playlist_id)
    os.makedirs(playlist_folder, exist_ok=True)

    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
        'ignoreerrors': True,
    }

    if USE_COOKIES:
        ydl_opts['cookiefile'] = COOKIES_PATH

    if format_type == "mp3":
        ydl_opts.update({
            'format': 'bestaudio/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'postprocessors': [{
                'key': 'FFmpegExtractAudio',
                'preferredcodec': 'mp3',
                'preferredquality': '192',
            }],
        })
    else:
        ydl_opts.update({
            'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'outtmpl': f'{playlist_folder}/%(playlist_index)s - %(title)s.%(ext)s',
            'merge_output_format': 'mp4',
        })

    zip_path = None
    try:
        with YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])

        if not os.listdir(playlist_folder):
            shutil.rmtree(playlist_folder)
            return {"error": "No se pudo descargar ningún video de la playlist"}

        zip_path = f"{output_folder}/{playlist_id}"
        shutil.make_archive(zip_path, 'zip', playlist_folder)
        shutil.rmtree(playlist_folder)

        def cleanup():
            try:
                if os.path.exists(f"{zip_path}.zip"):
                    os.remove(f"{zip_path}.zip")
            except Exception:
                pass

        return FileResponse(
            path=f"{zip_path}.zip",
            filename="playlist_download.zip",
            media_type="application/zip",
            background=cleanup
        )
    except Exception as e:
        if os.path.exists(playlist_folder):
            shutil.rmtree(playlist_folder)
        if zip_path and os.path.exists(f"{zip_path}.zip"):
            os.remove(f"{zip_path}.zip")

        error_msg = str(e)
        if "playlist" in error_msg.lower() and "not" in error_msg.lower():
            error_msg = "La URL no parece ser una playlist válida"

        return {"error": error_msg}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)